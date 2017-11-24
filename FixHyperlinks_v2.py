from openpyxl import load_workbook
import os
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# number of sheets, columns and rows to be examined and fixed (should be set by the user)
sheets = 1
columns = 8
rows = 200

Tk().withdraw()

file = askopenfilename() # show an "Open" dialog box and return the path to the selected file


file_directory = os.path.dirname(file) # extract file directory
file_full_name = os.path.basename(file) # extract file name with extension
file_name = os.path.splitext(file_full_name)[0] # extract file name without extenion
file_extension = os.path.splitext(file_full_name)[1] # extract file extension

wb = load_workbook(file) # read the xlsx file (should be in the working directory)

# defining function to loop trough the cells in the sheet and fix the hyperlinks
def sheetloop(columns, rows):
    for c in range(1,columns):
        for i in range(1,rows):
            link = sheet.cell(column=c, row=i).hyperlink # select the hyperlink
            if link != None: # change only if there hyperlink
                linktarget = link.target
                if "Excel" in linktarget: # change only if there is "Excel" in the hyperlink
                    exlplace = linktarget.index("Excel") # find the place of "Excel" in the string
                    newlink = linktarget[exlplace + 6:] # make new link from the end of the old one
                    sheet.cell(column=c, row=i).hyperlink = newlink # set the new link
                else: continue
            else: continue

wsnames = wb.get_sheet_names() # list of sheet names

if sheets > len(wsnames):
    sheets = len(wsnames)

for sh in range(0,sheets): # loop through the sheets
    sheet = wb[wsnames[sh]] # make object from the sheet
    sheetloop(columns, rows)

wb.save(file_directory+"/"+file_name+"_fixed"+file_extension)


